from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, ROUND_CEILING
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[2]
DATA_DIR = BASE_DIR / "data" / "calculation_sources"


def source_file(exact_name: str, portable_pattern: str) -> Path:
    """Resolve release assets even if ZIP extraction changed Unicode form.

    Windows-created archives may store Hungarian filenames in a normalization
    form that is rendered differently on Linux. The ASCII portion of each
    canonical filename is stable and the match must remain unique.
    """

    exact = DATA_DIR / exact_name
    if exact.exists():
        return exact
    matches = sorted(DATA_DIR.glob(portable_pattern))
    if len(matches) != 1:
        raise FileNotFoundError(
            f"Exactly one calculation source must match {portable_pattern!r}; found {len(matches)}."
        )
    return matches[0]


WEB_PRICES_FILE = source_file(
    "Kalkuláció_oldalakhoz_frissített_minden_weboldal_2026-07.xlsx",
    "*oldalakhoz*weboldal_2026-07.xlsx",
)
INTERNAL_MODEL_FILE = DATA_DIR / "Imperial_100m2_Technologia_Keszultseg_Armodell_2026_07.xlsx"
ARTUKOR_FILE = DATA_DIR / "Generalkivitelezo_ArTukor_Munkadij_Anyag_2026_07.xlsx"

BRAND_SHEET = {
    "imperial": "Imperial kalkulációs",
    "imperial holding": "Imperial kalkulációs",
    "bautica": "Bautica",
    "prefab": "Prefab",
    "danish fabrik": "Danish Fabrik",
    "baufreund": "BauFreund",
}

TECH_ALIASES = {
    "favázas könnyűszerkezet": "Danish Fabrik",
    "favazas konnyuszerkezet": "Danish Fabrik",
    "danish fabrik / faváz": "Danish Fabrik",
    "danish fabrik / készház": "Danish Fabrik",
    "danish fabrik": "Danish Fabrik",
    "tégla": "Tégla",
    "tegla": "Tégla",
    "leier falpanel": "Leier",
    "leier": "Leier",
    "liapor panel": "Liapor",
    "liapor": "Liapor",
    "acél": "Acél",
    "acel": "Acél",
    "acélvázas": "Acél",
    "clt": "CLT",
    "ytong": "Ytong",
    "sip": "Sip",
}

INTERNAL_TECH = {
    "Danish Fabrik": "Favázas könnyűszerkezet",
    "Tégla": "Tégla",
    "Leier": "Leier falpanel",
    "Liapor": "Liapor falpanel",
    "Acél": "Favázas könnyűszerkezet",
    "CLT": "Favázas könnyűszerkezet",
    "Ytong": "Tégla",
    "Sip": "Favázas könnyűszerkezet",
}


def D(value: Any) -> Decimal:
    return Decimal(str(value or 0))


def norm(value: str) -> str:
    return " ".join((value or "").strip().lower().split())


def round_up(value: Decimal, step: Decimal = Decimal("5000")) -> Decimal:
    if value <= 0:
        return Decimal("0")
    return (value / step).to_integral_value(rounding=ROUND_CEILING) * step


@dataclass(frozen=True)
class PriceSourceSnapshot:
    effective_date: str
    public_source: str
    internal_source: str
    artukor_source: str


class PricingRepository:
    def __init__(self) -> None:
        self._brand_prices: dict[str, dict[str, Decimal]] | None = None
        self._factors: dict[tuple[str, str], Decimal] | None = None
        self._internal: dict[tuple[str, str, str], dict[str, Decimal | str]] | None = None
        self._renovation_catalog: list[dict[str, Any]] | None = None

    def _load_public(self) -> None:
        if self._brand_prices is not None:
            return
        wb = load_workbook(WEB_PRICES_FILE, data_only=True, read_only=False)
        brand_prices: dict[str, dict[str, Decimal]] = {}
        for brand_key, sheet_name in BRAND_SHEET.items():
            ws = wb[sheet_name]
            prices: dict[str, Decimal] = {}
            for row in ws.iter_rows(values_only=True):
                tech, price = row[0], row[1]
                if isinstance(tech, str) and isinstance(price, (int, float)):
                    prices[TECH_ALIASES.get(norm(tech), tech.strip())] = D(price)
            brand_prices[brand_key] = prices
        factors: dict[tuple[str, str], Decimal] = {}
        ws = wb["Átadási tényezők"]
        for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
            completion = str(row[0])
            for idx, package in enumerate(("Alap", "Közép", "Prémium"), start=1):
                factors[(completion, package)] = D(row[idx])
        self._brand_prices = brand_prices
        self._factors = factors

    def _load_internal(self) -> None:
        if self._internal is not None:
            return
        wb = load_workbook(INTERNAL_MODEL_FILE, data_only=True, read_only=False)
        ws = wb["Haz_Ar_Matrix"]
        rows: dict[tuple[str, str, str], dict[str, Decimal | str]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            key = (str(row[0]), str(row[1]), str(row[2]))
            rows[key] = {
                "cash_cost_m2": D(row[8]),
                "minimum_m2": D(row[10]),
                "optimum_m2": D(row[12]),
                "market_ceiling_m2": D(row[14]),
                "recommended_m2": D(row[18]),
                "market_margin": D(row[16]),
                "optimum_reachable": str(row[17]),
            }
        self._internal = rows

    def brand_catalog(self) -> dict[str, list[str]]:
        self._load_public()
        assert self._brand_prices is not None
        return {brand: sorted(prices) for brand, prices in self._brand_prices.items() if brand in {"imperial", "bautica", "prefab", "danish fabrik", "baufreund"}}

    def calculate_new_build(
        self,
        brand: str,
        technology: str,
        completion_level: str,
        package: str,
        gross_area_m2: Decimal,
        vat_rate: Decimal = Decimal("0.05"),
        include_internal: bool = False,
    ) -> dict[str, Any]:
        self._load_public()
        self._load_internal()
        assert self._brand_prices is not None and self._factors is not None and self._internal is not None
        brand_key = norm(brand)
        if brand_key not in self._brand_prices:
            raise ValueError("Ismeretlen márka.")
        tech_key = TECH_ALIASES.get(norm(technology), technology.strip())
        base_price = self._brand_prices[brand_key].get(tech_key)
        if base_price is None:
            raise ValueError("A kiválasztott technológia ehhez a márkához nincs jóváhagyott árforrással összekötve.")
        factor = self._factors.get((completion_level, package))
        if factor is None:
            raise ValueError("Ismeretlen készültségi szint vagy műszaki csomag.")
        if gross_area_m2 < Decimal("20") or gross_area_m2 > Decimal("500"):
            raise ValueError("A kalkulálható bruttó alapterület 20–500 m².")
        public_unit_gross = round_up(base_price * factor)
        public_total_gross = round_up(public_unit_gross * gross_area_m2, Decimal("100000"))
        public_total_net = (public_total_gross / (Decimal("1") + vat_rate)).quantize(Decimal("1"))
        result: dict[str, Any] = {
            "brand": brand,
            "technology": tech_key,
            "completion_level": completion_level,
            "package": package,
            "gross_area_m2": str(gross_area_m2),
            "vat_rate": str(vat_rate),
            "estimated_gross_unit_price_huf": int(public_unit_gross),
            "estimated_gross_total_huf": int(public_total_gross),
            "estimated_net_total_huf": int(public_total_net),
            "price_basis": "2026-07 jóváhagyott márkaár × készültségi/csomagtényező",
            "source_version": "Kalkuláció_oldalakhoz_frissített_minden_weboldal_2026-07",
            "assumptions": [
                "Tájékoztató előkalkuláció; telek-, terv- és műszaki pontosítás nélkül.",
                "A publikus eredmény nem tartalmaz belső önköltséget, fedezeti vagy kapacitásadatot.",
                "A szerződéses ár kizárólag jóváhagyott BuildConfig-verzióból adható ki.",
            ],
        }
        if include_internal:
            internal_tech = INTERNAL_TECH.get(tech_key)
            internal = self._internal.get((internal_tech, completion_level, package)) if internal_tech else None
            if internal:
                cash_cost = D(internal["cash_cost_m2"]) * gross_area_m2
                margin = Decimal("1") - (cash_cost / public_total_gross) if public_total_gross else Decimal("0")
                result["internal_control"] = {
                    "internal_technology": internal_tech,
                    "cash_cost_m2_huf": int(D(internal["cash_cost_m2"])),
                    "cash_cost_total_huf": int(cash_cost),
                    "minimum_price_m2_huf": int(D(internal["minimum_m2"])),
                    "optimum_price_m2_huf": int(D(internal["optimum_m2"])),
                    "market_ceiling_m2_huf": int(D(internal["market_ceiling_m2"])),
                    "calculated_cash_margin": str(margin.quantize(Decimal("0.0001"))),
                    "margin_gate": "pass" if margin >= Decimal("0.35") else "stop",
                    "note": "A belső értékek kizárólag jogosult felhasználónak jelenhetnek meg.",
                }
        return result

    def _load_renovation_catalog(self) -> None:
        if self._renovation_catalog is not None:
            return
        wb = load_workbook(ARTUKOR_FILE, data_only=True, read_only=False)
        catalog: list[dict[str, Any]] = []
        ws = wb["Artukor_Master"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not isinstance(row[7], (int, float)):
                continue
            catalog.append({
                "item_id": str(row[0]),
                "type": "labor",
                "trade": row[1],
                "group": row[2],
                "technology": row[3],
                "name": row[4],
                "unit": row[5],
                "net_unit_price_huf": int(row[7]),
                "data_quality": row[28],
                "note": row[29],
            })
        ws = wb["Anyag_Master"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not isinstance(row[25], (int, float)):
                continue
            catalog.append({
                "item_id": str(row[0]),
                "type": "material",
                "trade": row[1],
                "group": row[2],
                "technology": row[3],
                "name": row[4],
                "specification": row[5],
                "unit": row[6],
                "net_unit_price_huf": int(row[25]),
                "data_quality": row[41],
                "note": row[42],
            })
        self._renovation_catalog = catalog

    def renovation_catalog(self, query: str = "", limit: int = 50) -> list[dict[str, Any]]:
        self._load_renovation_catalog()
        assert self._renovation_catalog is not None
        q = norm(query)
        rows = self._renovation_catalog
        if q:
            rows = [r for r in rows if q in norm(" ".join(str(r.get(k) or "") for k in ("item_id", "trade", "group", "name", "specification")))]
        return rows[: max(1, min(limit, 200))]

    def calculate_renovation(self, lines: list[dict[str, Any]], vat_rate: Decimal = Decimal("0.27")) -> dict[str, Any]:
        self._load_renovation_catalog()
        assert self._renovation_catalog is not None
        by_id = {r["item_id"]: r for r in self._renovation_catalog}
        result_lines: list[dict[str, Any]] = []
        net_total = Decimal("0")
        for line in lines:
            item_id = str(line.get("item_id") or "")
            quantity = D(line.get("quantity"))
            if item_id not in by_id:
                raise ValueError(f"Ismeretlen Ártükör-tétel: {item_id}")
            if quantity <= 0:
                raise ValueError("Minden mennyiségnek pozitívnak kell lennie.")
            item = by_id[item_id]
            line_net = D(item["net_unit_price_huf"]) * quantity
            net_total += line_net
            result_lines.append({**item, "quantity": str(quantity), "net_line_total_huf": int(line_net)})
        gross_total = (net_total * (Decimal("1") + vat_rate)).quantize(Decimal("1"))
        survey_upper = (gross_total * Decimal("1.20")).quantize(Decimal("1"))
        return {
            "lines": result_lines,
            "vat_rate": str(vat_rate),
            "estimated_net_total_huf": int(net_total),
            "estimated_gross_total_huf": int(gross_total),
            "pre_survey_upper_range_huf": int(survey_upper),
            "source_version": "Generalkivitelezo_ArTukor_Munkadij_Anyag_2026_07",
            "assumptions": [
                "A mennyiségeket a felhasználó vagy műszaki előkészítő adta meg.",
                "A 20%-os felső sáv helyszíni feltárás előtti bizonytalansági tartalék, nem automatikus szerződéses felár.",
                "A tételes Ártükör nem adható hozzá automatikusan újépítési all-in önköltséghez.",
            ],
        }


pricing_repository = PricingRepository()
