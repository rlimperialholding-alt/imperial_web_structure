from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[2]
CATALOG_FILE = BASE_DIR / "data" / "calculation_sources" / "HouseMatch_catalog_score_v0.1.xlsx"


def D(value: Any) -> Decimal:
    return Decimal(str(value or 0))


def norm(value: str) -> str:
    return " ".join((value or "").strip().lower().split())


@dataclass(frozen=True)
class HouseProfile:
    budget_huf: Decimal
    target_area_m2: Decimal
    lifestyle: str | None = None
    allowed_brands: tuple[str, ...] = ()
    score_profile: str = "Kiegyensúlyozott"


class HouseMatchRepository:
    def __init__(self) -> None:
        self._catalog: list[dict[str, Any]] | None = None
        self._weights: dict[str, dict[str, Decimal]] | None = None

    def _load(self) -> None:
        if self._catalog is not None:
            return
        wb = load_workbook(CATALOG_FILE, data_only=True, read_only=False)
        ws = wb["Katalogus"]
        catalog: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            active = bool(row[8])
            area = D(row[4])
            price = D(row[3])
            lifestyles = [norm(x) for x in str(row[10] or "").split(",") if x.strip()]
            catalog.append({
                "house_id": str(row[0]),
                "brand": str(row[1]),
                "name": str(row[2]),
                "catalog_price_huf": int(price),
                "gross_area_m2": float(area),
                "price_per_m2_huf": int(D(row[5])),
                "rooms": row[6],
                "price_status": row[7],
                "active": active,
                "data_quality": row[9],
                "lifestyles": lifestyles,
                "source_type": row[11],
                "source_url": row[12],
                "verified_at": str(row[13] or ""),
                "note": row[14],
            })
        ws = wb["Pontozasi_modell"]
        weights: dict[str, dict[str, Decimal]] = {}
        for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
            weights[str(row[0])] = {
                "price": D(row[1]) / Decimal("100"),
                "area": D(row[2]) / Decimal("100"),
                "lifestyle": D(row[3]) / Decimal("100"),
                "brand": D(row[4]) / Decimal("100"),
            }
        self._catalog = catalog
        self._weights = weights

    def catalog(self, brand: str | None = None, active_only: bool = True) -> list[dict[str, Any]]:
        self._load()
        assert self._catalog is not None
        rows = self._catalog
        if active_only:
            rows = [r for r in rows if r["active"]]
        if brand:
            rows = [r for r in rows if norm(r["brand"]) == norm(brand)]
        return rows

    @staticmethod
    def _price_score(price: Decimal, budget: Decimal) -> Decimal:
        if budget <= 0:
            return Decimal("0")
        ratio = price / budget
        if ratio <= 1:
            return Decimal("1") - min(Decimal("0.35"), abs(Decimal("1") - ratio) * Decimal("0.35"))
        return max(Decimal("0"), Decimal("1") - (ratio - Decimal("1")) * Decimal("2.2"))

    @staticmethod
    def _area_score(area: Decimal, target: Decimal) -> Decimal:
        if target <= 0:
            return Decimal("0")
        return max(Decimal("0"), Decimal("1") - abs(area - target) / target * Decimal("1.8"))

    @staticmethod
    def _lifestyle_score(lifestyles: list[str], selected: str | None) -> Decimal:
        if not selected:
            return Decimal("0.75")
        return Decimal("1") if norm(selected) in lifestyles else Decimal("0.15")

    @staticmethod
    def _brand_score(brand: str, allowed: tuple[str, ...]) -> Decimal:
        if not allowed:
            return Decimal("1")
        return Decimal("1") if norm(brand) in {norm(x) for x in allowed} else Decimal("0")

    def match(
        self,
        profile: HouseProfile,
        limit: int = 6,
        catalog: list[dict[str, Any]] | None = None,
    ) -> list[dict[str, Any]]:
        self._load()
        assert self._catalog is not None and self._weights is not None
        weights = self._weights.get(profile.score_profile)
        if not weights:
            raise ValueError("Ismeretlen HouseMatch pontozási profil.")
        matches: list[dict[str, Any]] = []
        for house in catalog if catalog is not None else self._catalog:
            if not house["active"]:
                continue
            brand_score = self._brand_score(house["brand"], profile.allowed_brands)
            if brand_score == 0:
                continue
            price_score = self._price_score(D(house["catalog_price_huf"]), profile.budget_huf)
            area_score = self._area_score(D(house["gross_area_m2"]), profile.target_area_m2)
            lifestyle_score = self._lifestyle_score(house["lifestyles"], profile.lifestyle)
            total = (
                price_score * weights["price"]
                + area_score * weights["area"]
                + lifestyle_score * weights["lifestyle"]
                + brand_score * weights["brand"]
            ) * Decimal("100")
            reasons: list[str] = []
            if price_score >= Decimal("0.85"):
                reasons.append("jól illeszkedik a költségkerethez")
            elif price_score < Decimal("0.45"):
                reasons.append("a megadott kerethez képest jelentős kompromisszum")
            if area_score >= Decimal("0.85"):
                reasons.append("közel van a cél-alapterülethez")
            if profile.lifestyle and lifestyle_score == 1:
                reasons.append(f"illeszkedik ehhez az élethelyzethez: {profile.lifestyle}")
            matches.append({
                **house,
                "score": float(total.quantize(Decimal("0.1"))),
                "score_components": {
                    "price": float((price_score * Decimal("100")).quantize(Decimal("0.1"))),
                    "area": float((area_score * Decimal("100")).quantize(Decimal("0.1"))),
                    "lifestyle": float((lifestyle_score * Decimal("100")).quantize(Decimal("0.1"))),
                    "brand": float((brand_score * Decimal("100")).quantize(Decimal("0.1"))),
                },
                "reasons": reasons,
                "score_profile": profile.score_profile,
            })
        matches.sort(key=lambda r: (-r["score"], abs(D(r["catalog_price_huf"]) - profile.budget_huf)))
        return matches[: max(1, min(limit, 12))]


housematch_repository = HouseMatchRepository()
