from __future__ import annotations

import csv
import io
import json
from hashlib import sha256
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ALLOWED_EXTENSIONS = {".csv", ".json", ".xlsx", ".xlsm", ".txt"}
MAX_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_ROWS_PER_FILE = 20_000


def _clean_headers(values: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for idx, value in enumerate(values, start=1):
        base = str(value or f"column_{idx}").strip()
        if not base:
            base = f"column_{idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        result.append(base if count == 1 else f"{base}_{count}")
    return result


def parse_upload(filename: str, data: bytes) -> dict[str, Any]:
    if len(data) > MAX_UPLOAD_BYTES:
        raise ValueError("A feltöltött fájl legfeljebb 20 MB lehet.")
    suffix = Path(filename or "").suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise ValueError("Támogatott formátumok: CSV, JSON, XLSX, XLSM és TXT.")

    digest = sha256(data).hexdigest()
    if suffix == ".json":
        parsed = json.loads(data.decode("utf-8-sig"))
        if isinstance(parsed, list):
            json_records = [row for row in parsed if isinstance(row, dict)]
        elif isinstance(parsed, dict) and isinstance(parsed.get("records"), list):
            json_records = [row for row in parsed["records"] if isinstance(row, dict)]
        elif isinstance(parsed, dict):
            json_records = [parsed]
        else:
            raise ValueError("A JSON fájlnak objektumot vagy objektumlistát kell tartalmaznia.")
        return {"records": json_records[:MAX_ROWS_PER_FILE], "metadata": {"format": "json", "sha256": digest}}

    if suffix == ".csv":
        text = data.decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = None
        reader = csv.DictReader(io.StringIO(text), dialect=dialect) if dialect else csv.DictReader(io.StringIO(text), delimiter=";")
        csv_records = [dict(row) for _, row in zip(range(MAX_ROWS_PER_FILE), reader)]
        return {"records": csv_records, "metadata": {"format": "csv", "delimiter": dialect.delimiter if dialect else ";", "sha256": digest}}

    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        workbook_records: list[dict[str, Any]] = []
        sheet_stats: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            iterator = ws.iter_rows(values_only=True)
            header_row = next(iterator, None)
            if not header_row:
                continue
            headers = _clean_headers(list(header_row))
            count = 0
            for values in iterator:
                if count >= MAX_ROWS_PER_FILE:
                    break
                if not any(value not in (None, "") for value in values):
                    continue
                row = {headers[idx]: value for idx, value in enumerate(values[: len(headers)])}
                row["_source_sheet"] = ws.title
                workbook_records.append(row)
                count += 1
            sheet_stats.append({"sheet": ws.title, "rows": count})
            if len(workbook_records) >= MAX_ROWS_PER_FILE:
                break
        return {"records": workbook_records[:MAX_ROWS_PER_FILE], "metadata": {"format": "xlsx", "sheets": sheet_stats, "sha256": digest}}

    text = data.decode("utf-8-sig", errors="replace")
    return {"text": text[:500_000], "metadata": {"format": "text", "sha256": digest}}
